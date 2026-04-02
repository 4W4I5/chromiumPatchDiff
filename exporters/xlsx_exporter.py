from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _parse_iso_utc(value: str) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None

    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    return parsed.astimezone(timezone.utc)


def _days_since(value: str, now_utc: datetime) -> int | str:
    parsed = _parse_iso_utc(value)
    if parsed is None:
        return ""
    return max(0, (now_utc - parsed).days)


def _as_joined_lines(items: list[Any]) -> str:
    return "\n".join(str(item) for item in items if str(item).strip())


def _style_header_row(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize_columns(sheet, max_width: int = 80) -> None:
    for column_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(cell_value))

        desired = min(max_width, max(12, max_len + 2))
        sheet.column_dimensions[get_column_letter(column_idx)].width = desired


def write_enrichment_xlsx(result: dict[str, Any], output_path: str) -> None:
    workbook = Workbook()
    now_utc = datetime.now(timezone.utc)
    cves = result.get("cves", []) or []

    cve_sheet = workbook.active
    cve_sheet.title = "CVEs"
    cve_sheet.append(
        [
            "cve_id",
            "source",
            "title",
            "description",
            "published",
            "updated",
            "match_reason",
            "match_confidence",
            "has_commit",
            "commit_count",
            "max_commit_confidence",
            "references_count",
            "affected_versions_count",
            "days_since_published",
            "nvd_cvss_score",
            "nvd_cvss_vector",
            "nvd_severity",
            "nvd_weaknesses",
            "nvd_cpes",
            "references",
            "affected_versions",
            "commits",
        ]
    )

    for cve in cves:
        commits = cve.get("commits", []) or []
        references = cve.get("references", []) or []
        affected_versions = cve.get("affected_versions", []) or []
        nvd = cve.get("nvd") or {}
        commit_confidences = [float(item.get("confidence", 0.0) or 0.0) for item in commits if isinstance(item, dict)]

        cve_sheet.append(
            [
                cve.get("cve_id", ""),
                cve.get("source", ""),
                cve.get("title", ""),
                cve.get("description", ""),
                cve.get("published", ""),
                cve.get("updated", ""),
                cve.get("match_reason", ""),
                cve.get("match_confidence", 0),
                bool(commits),
                len(commits),
                max(commit_confidences) if commit_confidences else 0.0,
                len(references),
                len(affected_versions),
                _days_since(str(cve.get("published", "") or ""), now_utc),
                nvd.get("cvss_score", "") if isinstance(nvd, dict) else "",
                nvd.get("cvss_vector", "") if isinstance(nvd, dict) else "",
                nvd.get("severity", "") if isinstance(nvd, dict) else "",
                _as_joined_lines(nvd.get("weaknesses", []) if isinstance(nvd, dict) else []),
                _as_joined_lines(nvd.get("cpes", []) if isinstance(nvd, dict) else []),
                _as_joined_lines(references),
                _as_joined_lines(affected_versions),
                _as_joined_lines([f"{item.get('sha', '')} | {item.get('title', '')}" for item in commits if isinstance(item, dict)]),
            ]
        )

    commits_sheet = workbook.create_sheet(title="Commits")
    commits_sheet.append(["cve_id", "sha", "title", "url", "author", "date", "confidence", "source"])
    for cve in cves:
        cve_id = cve.get("cve_id", "")
        for commit in cve.get("commits", []) or []:
            if not isinstance(commit, dict):
                continue
            commits_sheet.append(
                [
                    cve_id,
                    commit.get("sha", ""),
                    commit.get("title", ""),
                    commit.get("url", ""),
                    commit.get("author", ""),
                    commit.get("date", ""),
                    commit.get("confidence", 0),
                    commit.get("source", ""),
                ]
            )

    references_sheet = workbook.create_sheet(title="References")
    references_sheet.append(["cve_id", "reference_url"])
    for cve in cves:
        cve_id = cve.get("cve_id", "")
        for reference in cve.get("references", []) or []:
            references_sheet.append([cve_id, reference])

    warnings_sheet = workbook.create_sheet(title="Warnings")
    warnings_sheet.append(["warning"])
    for warning in result.get("warnings", []) or []:
        warnings_sheet.append([warning])

    metadata_sheet = workbook.create_sheet(title="Metadata")
    metadata_sheet.append(["key", "value"])
    for key in (
        "input_version",
        "compare_base_version",
        "compare_commit_count",
        "source_mode",
        "selected_cve_source",
        "generated_at",
        "candidate_count",
        "matched_count",
    ):
        metadata_sheet.append([key, result.get(key, "")])

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_compare_xlsx(result: dict[str, Any], output_path: str) -> None:
    workbook = Workbook()

    metadata_sheet = workbook.active
    metadata_sheet.title = "Compare Metadata"
    metadata_sheet.append(["key", "value"])
    for key in (
        "task",
        "compare_repo",
        "compare_base_version",
        "compare_head_version",
        "compare_url",
        "compare_commit_count",
        "generated_at",
    ):
        metadata_sheet.append([key, result.get(key, "")])

    summary_sheet = workbook.create_sheet(title="Summary")
    summary_sheet.append(["metric", "value"])
    summary_sheet.append(["warnings_count", len(result.get("warnings", []) or [])])
    summary_sheet.append(["commit_count", int(result.get("compare_commit_count", 0) or 0)])

    links_sheet = workbook.create_sheet(title="Links")
    links_sheet.append(["label", "url"])
    compare_url = str(result.get("compare_url", "") or "")
    if compare_url:
        links_sheet.append(["github_compare", compare_url])
        compare_url_cell = links_sheet.cell(row=links_sheet.max_row, column=2)
        compare_url_cell.hyperlink = compare_url
        compare_url_cell.style = "Hyperlink"

    commits_sheet = workbook.create_sheet(title="Compare Commits")
    commits_sheet.append(["sha", "title", "url", "author", "date", "confidence", "source"])
    for commit in result.get("commits", []) or []:
        if not isinstance(commit, dict):
            continue

        commits_sheet.append(
            [
                commit.get("sha", ""),
                commit.get("title", ""),
                commit.get("url", ""),
                commit.get("author", ""),
                commit.get("date", ""),
                commit.get("confidence", 0),
                commit.get("source", ""),
            ]
        )

        commit_url = str(commit.get("url", "") or "")
        if commit_url:
            commit_url_cell = commits_sheet.cell(row=commits_sheet.max_row, column=3)
            commit_url_cell.hyperlink = commit_url
            commit_url_cell.style = "Hyperlink"

    warnings_sheet = workbook.create_sheet(title="Warnings")
    warnings_sheet.append(["warning"])
    for warning in result.get("warnings", []) or []:
        warnings_sheet.append([warning])

    for sheet in (metadata_sheet, summary_sheet, links_sheet, commits_sheet, warnings_sheet):
        sheet.freeze_panes = "A2"
        _style_header_row(sheet)
        _autosize_columns(sheet)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
